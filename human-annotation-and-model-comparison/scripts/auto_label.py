from pathlib import Path
from openpyxl import load_workbook


# =========================
# CONFIGURACIÓN
# =========================

INPUT_FILE = Path("sample_1000_sentences_for_manual_annotation_randomized.xlsx")
OUTPUT_FILE = Path("sample_1000_sentences_for_manual_annotation_randomized_auto_label.xlsx")

# Si quieres procesar la hoja activa, déjalo en None.
# Si tu hoja tiene un nombre concreto, escríbelo aquí.
SHEET_NAME = None

POS_COL = "F"
NEU_COL = "G"
NEG_COL = "H"
AUTO_LABEL_COL = "I"

FIRST_DATA_ROW = 2


# =========================
# FUNCIONES
# =========================

def to_float(value):
    """
    Convierte a número valores escritos con punto o coma decimal.
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    value = str(value).strip().replace(",", ".")

    if value == "":
        return None

    return float(value)


def get_auto_label(pos, neu, neg):
    """
    Devuelve la etiqueta con mayor probabilidad entre POS, NEU y NEG.
    """
    scores = {
        "POS": pos,
        "NEU": neu,
        "NEG": neg,
    }

    if any(value is None for value in scores.values()):
        return None

    return max(scores, key=scores.get)


# =========================
# LECTURA DEL EXCEL
# =========================

wb = load_workbook(INPUT_FILE)

if SHEET_NAME is None:
    ws = wb.active
else:
    ws = wb[SHEET_NAME]


# =========================
# CREAR AUTO_LABEL EN COLUMNA J
# =========================

ws[f"{AUTO_LABEL_COL}1"] = "auto_label"

for row in range(FIRST_DATA_ROW, ws.max_row + 1):
    pos = to_float(ws[f"{POS_COL}{row}"].value)
    neu = to_float(ws[f"{NEU_COL}{row}"].value)
    neg = to_float(ws[f"{NEG_COL}{row}"].value)

    auto_label = get_auto_label(pos, neu, neg)

    ws[f"{AUTO_LABEL_COL}{row}"] = auto_label


# =========================
# GUARDAR NUEVO ARCHIVO
# =========================

wb.save(OUTPUT_FILE)

print(f"Archivo creado correctamente: {OUTPUT_FILE}")
print(f"Hoja procesada: {ws.title}")
print("Columna J creada: auto_label")