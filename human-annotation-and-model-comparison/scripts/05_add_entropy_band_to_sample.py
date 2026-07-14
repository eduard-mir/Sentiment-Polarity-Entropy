from pathlib import Path
from openpyxl import load_workbook
import csv


# =========================
# CONFIGURACIÓN
# =========================

SAMPLE_FILE = Path("sample_1000_sentences_for_manual_annotation_randomized_auto_label_gold.xlsx")
SELECTED_ADJECTIVES_FILE = Path("selected_200_adjectives.csv")

OUTPUT_FILE = Path("sample_1000_sentences_for_manual_annotation_randomized_auto_label_gold_entropy_band.xlsx")

SHEET_NAME = None  # None = usa la hoja activa

FIRST_DATA_ROW = 2
LAST_DATA_ROW = 960

# Archivo principal
SAMPLE_WORD_COL = "C"
OUTPUT_ENTROPY_BAND_COL = "Q"

# Archivo selected_200_adjectives
# Usamos índices porque es CSV:
# columna B = índice 1
# columna I = índice 8
SELECTED_WORD_INDEX = 1
SELECTED_ENTROPY_BAND_INDEX = 8

VALID_ENTROPY_BANDS = {"low", "mid", "high"}


# =========================
# FUNCIONES
# =========================

def normalize_word(value):
    """
    Normaliza la palabra para hacer el cruce entre archivos.
    """
    if value is None:
        return ""
    return str(value).strip().lower()


def normalize_entropy_band(value):
    """
    Normaliza entropy_band a low / mid / high.
    """
    if value is None:
        return None

    value = str(value).strip().lower()

    if value in VALID_ENTROPY_BANDS:
        return value

    return None


def detect_delimiter(path):
    """
    Detecta si el CSV usa punto y coma o coma.
    """
    with open(path, mode="r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(2048)
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        return dialect.delimiter


def load_entropy_band_mapping(path):
    """
    Carga un diccionario:
        palabra -> entropy_band

    desde selected_200_adjectives.csv.
    """

    delimiter = detect_delimiter(path)

    mapping = {}

    with open(path, mode="r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)

        rows = list(reader)

    # Saltamos la primera fila porque contiene encabezados
    for row_number, row in enumerate(rows[1:], start=2):

        if len(row) <= max(SELECTED_WORD_INDEX, SELECTED_ENTROPY_BAND_INDEX):
            raise ValueError(
                f"La fila {row_number} de {path} no tiene suficientes columnas: {row}"
            )

        word = normalize_word(row[SELECTED_WORD_INDEX])
        entropy_band = normalize_entropy_band(row[SELECTED_ENTROPY_BAND_INDEX])

        if not word:
            continue

        if entropy_band is None:
            raise ValueError(
                f"Valor entropy_band no válido en fila {row_number}: "
                f"{row[SELECTED_ENTROPY_BAND_INDEX]}"
            )

        mapping[word] = entropy_band

    return mapping


# =========================
# CARGAR MAPA PALABRA -> ENTROPY_BAND
# =========================

entropy_band_by_word = load_entropy_band_mapping(SELECTED_ADJECTIVES_FILE)

print(f"Palabras cargadas desde {SELECTED_ADJECTIVES_FILE}: {len(entropy_band_by_word)}")


# =========================
# LEER EXCEL PRINCIPAL
# =========================

wb = load_workbook(SAMPLE_FILE)

if SHEET_NAME is None:
    ws = wb.active
else:
    ws = wb[SHEET_NAME]


# =========================
# AÑADIR ENTROPY_BAND EN COLUMNA Q
# =========================

ws[f"{OUTPUT_ENTROPY_BAND_COL}1"] = "entropy_band"

matched_count = 0
missing_words = []

for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):

    word_raw = ws[f"{SAMPLE_WORD_COL}{row}"].value
    word_key = normalize_word(word_raw)

    entropy_band = entropy_band_by_word.get(word_key)

    if entropy_band is None:
        missing_words.append({
            "excel_row": row,
            "word": word_raw,
        })
        ws[f"{OUTPUT_ENTROPY_BAND_COL}{row}"] = None
    else:
        ws[f"{OUTPUT_ENTROPY_BAND_COL}{row}"] = entropy_band
        matched_count += 1


# =========================
# CONTROL DE ERRORES
# =========================

if missing_words:
    print()
    print("ADVERTENCIA: hay palabras sin entropy_band.")
    print("Primeros casos:")

    for item in missing_words[:20]:
        print(f"Fila {item['excel_row']}: {item['word']}")

    raise ValueError(
        f"No se encontró entropy_band para {len(missing_words)} filas. "
        "Revisa que las palabras de la columna C coincidan con las de selected_200_adjectives."
    )


# =========================
# GUARDAR NUEVO EXCEL
# =========================

wb.save(OUTPUT_FILE)

print()
print("ENTROPY_BAND AÑADIDO CORRECTAMENTE")
print("==================================")
print(f"Archivo de entrada: {SAMPLE_FILE}")
print(f"Archivo de referencia: {SELECTED_ADJECTIVES_FILE}")
print(f"Archivo de salida: {OUTPUT_FILE}")
print(f"Hoja procesada: {ws.title}")
print(f"Filas procesadas: {FIRST_DATA_ROW}-{LAST_DATA_ROW}")
print(f"Filas con entropy_band asignado: {matched_count}")
print(f"Columna creada: {OUTPUT_ENTROPY_BAND_COL} = entropy_band")