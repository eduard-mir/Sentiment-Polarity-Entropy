from pathlib import Path
from openpyxl import load_workbook
import csv


# =========================
# CONFIGURACIÓN
# =========================

INPUT_FILE = Path("sample_1000_sentences_for_manual_annotation_randomized_auto_label.xlsx")

OUTPUT_SUMMARY_CSV = Path("human_raw_agreement_summary.csv")
OUTPUT_CASES_CSV = Path("human_raw_agreement_cases.csv")

SHEET_NAME = None  # None = usa la hoja activa

FIRST_DATA_ROW = 2
LAST_DATA_ROW = 960

ANNOTATOR_1_COL = "J"
ANNOTATOR_2_COL = "K"
ANNOTATOR_3_COL = "L"


# =========================
# FUNCIONES
# =========================

def classify_agreement(label1, label2, label3):
    """
    Clasifica el acuerdo entre tres anotadores.
    """

    labels = {label1, label2, label3}

    if len(labels) == 1:
        return "full_agreement"

    if len(labels) == 2:
        return "partial_agreement"

    return "no_majority"


def percentage(count, total):
    return (count / total) * 100 if total > 0 else 0


# =========================
# LECTURA DEL EXCEL
# =========================

wb = load_workbook(INPUT_FILE, data_only=True)

if SHEET_NAME is None:
    ws = wb.active
else:
    ws = wb[SHEET_NAME]


# =========================
# CÁLCULO DEL ACUERDO
# =========================

counts = {
    "full_agreement": 0,
    "partial_agreement": 0,
    "no_majority": 0,
}

case_rows = []

for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):

    ann1 = ws[f"{ANNOTATOR_1_COL}{row}"].value
    ann2 = ws[f"{ANNOTATOR_2_COL}{row}"].value
    ann3 = ws[f"{ANNOTATOR_3_COL}{row}"].value

    agreement_type = classify_agreement(ann1, ann2, ann3)

    counts[agreement_type] += 1

    case_rows.append({
        "excel_row": row,
        "human_annotation_1": ann1,
        "human_annotation_2": ann2,
        "human_annotation_3": ann3,
        "agreement_type": agreement_type,
    })


total_cases = len(case_rows)


# =========================
# RESUMEN
# =========================

summary_rows = [
    {
        "agreement_type": "full_agreement",
        "description": "Los 3 anotadores coinciden",
        "count": counts["full_agreement"],
        "percentage": percentage(counts["full_agreement"], total_cases),
    },
    {
        "agreement_type": "partial_agreement",
        "description": "2 de 3 anotadores coinciden",
        "count": counts["partial_agreement"],
        "percentage": percentage(counts["partial_agreement"], total_cases),
    },
    {
        "agreement_type": "no_majority",
        "description": "Los 3 anotadores difieren",
        "count": counts["no_majority"],
        "percentage": percentage(counts["no_majority"], total_cases),
    },
]


# =========================
# IMPRIMIR RESULTADOS
# =========================

print("ACUERDO BRUTO ENTRE ANOTADORES HUMANOS")
print("=====================================")
print(f"Archivo analizado: {INPUT_FILE}")
print(f"Hoja analizada: {ws.title}")
print(f"Filas analizadas: {FIRST_DATA_ROW}-{LAST_DATA_ROW}")
print(f"Total de casos: {total_cases}")
print()

for item in summary_rows:
    print(
        f"{item['description']}: "
        f"{item['count']} casos "
        f"({item['percentage']:.2f}%)"
    )


# =========================
# GUARDAR RESUMEN
# =========================

with open(OUTPUT_SUMMARY_CSV, mode="w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(
        f,
        fieldnames=[
            "agreement_type",
            "description",
            "count",
            "percentage",
        ],
        delimiter=";"
    )
    writer.writeheader()
    writer.writerows(summary_rows)


# =========================
# GUARDAR RESULTADO POR FILA
# =========================

with open(OUTPUT_CASES_CSV, mode="w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(
        f,
        fieldnames=[
            "excel_row",
            "human_annotation_1",
            "human_annotation_2",
            "human_annotation_3",
            "agreement_type",
        ],
        delimiter=";"
    )
    writer.writeheader()
    writer.writerows(case_rows)


print()
print(f"Resumen guardado en: {OUTPUT_SUMMARY_CSV}")
print(f"Detalle por caso guardado en: {OUTPUT_CASES_CSV}")