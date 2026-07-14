from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# =========================
# CONFIGURACIÓN
# =========================

INPUT_FILE = Path("sample_1000_sentences_for_manual_annotation_randomized_auto_label_gold_entropy_band.xlsx")
OUTPUT_FILE = Path("auto_vs_gold_precision_recall_f1.xlsx")

SHEET_NAME = None  # None = usa la hoja activa

FIRST_DATA_ROW = 2
LAST_DATA_ROW = 960

AUTO_LABEL_COL = "I"
GOLD_LABEL_COL = "O"

LABELS = ["NEG", "NEU", "POS"]


# =========================
# FUNCIONES
# =========================

def clean_label(value):
    if value is None:
        return None

    value = str(value).strip().upper()

    if value in LABELS:
        return value

    return None


def safe_divide(numerator, denominator):
    if denominator == 0:
        return 0.0
    return numerator / denominator


def compute_metrics_for_label(gold_labels, auto_labels, target_label):
    """
    Calcula TP, FP, FN, TN, precision, recall y F1 para una etiqueta.
    """

    tp = fp = fn = tn = 0

    for gold, auto in zip(gold_labels, auto_labels):

        if gold == target_label and auto == target_label:
            tp += 1
        elif gold != target_label and auto == target_label:
            fp += 1
        elif gold == target_label and auto != target_label:
            fn += 1
        else:
            tn += 1

    precision = safe_divide(tp, tp + fp)
    recall = safe_divide(tp, tp + fn)

    if precision + recall == 0:
        f1 = 0.0
    else:
        f1 = 2 * (precision * recall) / (precision + recall)

    return {
        "label": target_label,
        "true_positives": tp,
        "false_positives": fp,
        "false_negatives": fn,
        "true_negatives": tn,
        "gold_support": tp + fn,
        "auto_predicted": tp + fp,
        "precision": precision,
        "recall": recall,
        "f1": f1,
    }


def build_confusion_matrix(gold_labels, auto_labels):
    """
    Matriz de confusión:
    filas = Gold Human Label
    columnas = auto_label
    """

    matrix = {
        gold_label: {auto_label: 0 for auto_label in LABELS}
        for gold_label in LABELS
    }

    for gold, auto in zip(gold_labels, auto_labels):
        matrix[gold][auto] += 1

    return matrix


def style_header(row):
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")


def apply_basic_style(ws):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))

        ws.column_dimensions[column_letter].width = min(max_length + 3, 28)


# =========================
# LECTURA DEL EXCEL
# =========================

wb_input = load_workbook(INPUT_FILE, data_only=True)

if SHEET_NAME is None:
    ws_input = wb_input.active
else:
    ws_input = wb_input[SHEET_NAME]


# =========================
# EXTRAER ETIQUETAS
# =========================

gold_labels = []
auto_labels = []
invalid_rows = []

for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):

    auto_label = clean_label(ws_input[f"{AUTO_LABEL_COL}{row}"].value)
    gold_label = clean_label(ws_input[f"{GOLD_LABEL_COL}{row}"].value)

    if auto_label is None or gold_label is None:
        invalid_rows.append({
            "excel_row": row,
            "auto_label": ws_input[f"{AUTO_LABEL_COL}{row}"].value,
            "gold_label": ws_input[f"{GOLD_LABEL_COL}{row}"].value,
        })
        continue

    auto_labels.append(auto_label)
    gold_labels.append(gold_label)


valid_total = len(gold_labels)
correct_total = sum(
    1 for gold, auto in zip(gold_labels, auto_labels)
    if gold == auto
)

accuracy = safe_divide(correct_total, valid_total)


# =========================
# CALCULAR MÉTRICAS
# =========================

metrics = [
    compute_metrics_for_label(gold_labels, auto_labels, label)
    for label in LABELS
]

macro_precision = sum(item["precision"] for item in metrics) / len(LABELS)
macro_recall = sum(item["recall"] for item in metrics) / len(LABELS)
macro_f1 = sum(item["f1"] for item in metrics) / len(LABELS)

confusion_matrix = build_confusion_matrix(gold_labels, auto_labels)


# =========================
# CREAR EXCEL DE SALIDA
# =========================

wb_output = Workbook()

# Hoja 1: métricas por clase
ws_metrics = wb_output.active
ws_metrics.title = "per_label_metrics"

headers = [
    "label",
    "gold_support",
    "auto_predicted",
    "true_positives",
    "false_positives",
    "false_negatives",
    "true_negatives",
    "precision",
    "recall",
    "f1",
]

ws_metrics.append(headers)
style_header(ws_metrics[1])

for item in metrics:
    ws_metrics.append([
        item["label"],
        item["gold_support"],
        item["auto_predicted"],
        item["true_positives"],
        item["false_positives"],
        item["false_negatives"],
        item["true_negatives"],
        item["precision"],
        item["recall"],
        item["f1"],
    ])

ws_metrics.append([])
ws_metrics.append([
    "MACRO_AVERAGE",
    "",
    "",
    "",
    "",
    "",
    "",
    macro_precision,
    macro_recall,
    macro_f1,
])

for row in range(2, ws_metrics.max_row + 1):
    for col in ["H", "I", "J"]:
        ws_metrics[f"{col}{row}"].number_format = "0.0000"

apply_basic_style(ws_metrics)


# Hoja 2: matriz de confusión
ws_confusion = wb_output.create_sheet("confusion_matrix")

ws_confusion.append(["Gold \\ Auto"] + LABELS)
style_header(ws_confusion[1])

for gold_label in LABELS:
    ws_confusion.append(
        [gold_label] + [confusion_matrix[gold_label][auto_label] for auto_label in LABELS]
    )

apply_basic_style(ws_confusion)


# Hoja 3: resumen global
ws_summary = wb_output.create_sheet("summary")

summary_rows = [
    ["input_file", str(INPUT_FILE)],
    ["input_sheet", ws_input.title],
    ["rows_checked", LAST_DATA_ROW - FIRST_DATA_ROW + 1],
    ["valid_cases", valid_total],
    ["invalid_or_missing_cases", len(invalid_rows)],
    ["correct_cases", correct_total],
    ["incorrect_cases", valid_total - correct_total],
    ["accuracy", accuracy],
    ["macro_precision", macro_precision],
    ["macro_recall", macro_recall],
    ["macro_f1", macro_f1],
]

for row in summary_rows:
    ws_summary.append(row)

ws_summary["A1"].font = Font(bold=True)
ws_summary["B1"].font = Font(bold=True)

for row in range(8, 12):
    ws_summary[f"B{row}"].number_format = "0.0000"

apply_basic_style(ws_summary)


# Hoja 4: filas inválidas, si las hay
ws_invalid = wb_output.create_sheet("invalid_rows")

ws_invalid.append(["excel_row", "auto_label_raw", "gold_label_raw"])
style_header(ws_invalid[1])

for item in invalid_rows:
    ws_invalid.append([
        item["excel_row"],
        item["auto_label"],
        item["gold_label"],
    ])

apply_basic_style(ws_invalid)


# =========================
# GUARDAR EXCEL
# =========================

wb_output.save(OUTPUT_FILE)


# =========================
# RESULTADOS EN CONSOLA
# =========================

print("PRECISION, RECALL, F1 Y MACRO-F1")
print("================================")
print(f"Archivo analizado: {INPUT_FILE}")
print(f"Hoja analizada: {ws_input.title}")
print(f"Filas analizadas: {FIRST_DATA_ROW}-{LAST_DATA_ROW}")
print(f"Casos válidos: {valid_total}")
print(f"Casos correctos: {correct_total}")
print(f"Accuracy: {accuracy:.4f}")
print()

for item in metrics:
    print(
        f"{item['label']}: "
        f"Precision={item['precision']:.4f}, "
        f"Recall={item['recall']:.4f}, "
        f"F1={item['f1']:.4f}"
    )

print()
print(f"Macro-Precision: {macro_precision:.4f}")
print(f"Macro-Recall: {macro_recall:.4f}")
print(f"Macro-F1: {macro_f1:.4f}")
print()
print(f"Archivo creado: {OUTPUT_FILE}")