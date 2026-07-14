from pathlib import Path
from collections import Counter
from openpyxl import load_workbook
import csv


# =========================
# CONFIGURACIÓN
# =========================

INPUT_FILE = Path("sample_1000_sentences_for_manual_annotation_randomized_auto_label.xlsx")
OUTPUT_FILE = Path("sample_1000_sentences_for_manual_annotation_randomized_auto_label_gold.xlsx")

OUTPUT_SUMMARY_CSV = Path("gold_human_label_summary.csv")

SHEET_NAME = None  # None = usa la hoja activa

FIRST_DATA_ROW = 2
LAST_DATA_ROW = 960

ANNOTATOR_1_COL = "J"
ANNOTATOR_2_COL = "K"
ANNOTATOR_3_COL = "L"
AGREEMENT_TYPE_COL = "M"
ADJUDICATION_COL = "N"
GOLD_LABEL_COL = "O"

VALID_LABELS = {"NEG", "NEU", "POS"}


# =========================
# FUNCIONES AUXILIARES
# =========================

def clean_label(value):
    """
    Limpia una etiqueta POS / NEU / NEG.
    """
    if value is None:
        return None
    value = str(value).strip().upper()
    return value if value in VALID_LABELS else None


def clean_agreement_type(value):
    """
    Limpia el valor de Agreement_Type.
    """
    if value is None:
        return None
    return str(value).strip().lower()


def majority_label(labels):
    """
    Devuelve la etiqueta mayoritaria entre tres anotaciones.
    Si no hay mayoría, devuelve None.
    """
    counts = Counter(labels)
    most_common = counts.most_common()

    if len(most_common) == 0:
        return None

    top_label, top_count = most_common[0]

    if top_count >= 2:
        return top_label

    return None


def create_gold_label(ann1, ann2, ann3, agreement_type, adjudication):
    """
    Crea la Gold Human Label según el tipo de acuerdo.
    """

    labels = [ann1, ann2, ann3]

    if agreement_type == "full_agreement":
        # Los tres anotadores coinciden.
        return ann1

    if agreement_type == "partial_agreement":
        # Dos de tres anotadores coinciden.
        return majority_label(labels)

    if agreement_type == "no_majority":
        # Se usa la adjudicación adicional de la columna N.
        return adjudication

    return None


# =========================
# LECTURA DEL EXCEL
# =========================

wb = load_workbook(INPUT_FILE)

if SHEET_NAME is None:
    ws = wb.active
else:
    ws = wb[SHEET_NAME]


# =========================
# CREACIÓN DE GOLD HUMAN LABEL
# =========================

ws[f"{GOLD_LABEL_COL}1"] = "Gold Human Label"

summary_counts = {
    "full_agreement": 0,
    "partial_agreement": 0,
    "no_majority": 0,
    "invalid_or_missing": 0,
}

problem_rows = []

for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):

    ann1 = clean_label(ws[f"{ANNOTATOR_1_COL}{row}"].value)
    ann2 = clean_label(ws[f"{ANNOTATOR_2_COL}{row}"].value)
    ann3 = clean_label(ws[f"{ANNOTATOR_3_COL}{row}"].value)

    agreement_type = clean_agreement_type(ws[f"{AGREEMENT_TYPE_COL}{row}"].value)
    adjudication = clean_label(ws[f"{ADJUDICATION_COL}{row}"].value)

    gold_label = create_gold_label(
        ann1=ann1,
        ann2=ann2,
        ann3=ann3,
        agreement_type=agreement_type,
        adjudication=adjudication
    )

    if gold_label is None:
        summary_counts["invalid_or_missing"] += 1
        problem_rows.append({
            "excel_row": row,
            "human_annotation_1": ann1,
            "human_annotation_2": ann2,
            "human_annotation_3": ann3,
            "agreement_type": agreement_type,
            "human_annotation_4_5": adjudication,
        })
    else:
        summary_counts[agreement_type] += 1

    ws[f"{GOLD_LABEL_COL}{row}"] = gold_label


# =========================
# GUARDAR NUEVO EXCEL
# =========================

wb.save(OUTPUT_FILE)


# =========================
# GUARDAR RESUMEN METODOLÓGICO
# =========================

total_rows = LAST_DATA_ROW - FIRST_DATA_ROW + 1

summary_rows = [
    {
        "category": "full_agreement",
        "description": "Gold label taken from the common label in Human Annotation 1, 2 and 3",
        "count": summary_counts["full_agreement"],
        "percentage": (summary_counts["full_agreement"] / total_rows) * 100,
    },
    {
        "category": "partial_agreement",
        "description": "Gold label taken from the majority label among Human Annotation 1, 2 and 3",
        "count": summary_counts["partial_agreement"],
        "percentage": (summary_counts["partial_agreement"] / total_rows) * 100,
    },
    {
        "category": "no_majority",
        "description": "Gold label taken from Human Annotation 4 & 5",
        "count": summary_counts["no_majority"],
        "percentage": (summary_counts["no_majority"] / total_rows) * 100,
    },
    {
        "category": "invalid_or_missing",
        "description": "Rows with missing or invalid labels",
        "count": summary_counts["invalid_or_missing"],
        "percentage": (summary_counts["invalid_or_missing"] / total_rows) * 100,
    },
]

with open(OUTPUT_SUMMARY_CSV, mode="w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(
        f,
        fieldnames=["category", "description", "count", "percentage"],
        delimiter=";"
    )
    writer.writeheader()
    writer.writerows(summary_rows)


# =========================
# CONTROL FINAL
# =========================

print("GOLD HUMAN LABEL CREADA")
print("=======================")
print(f"Archivo de entrada: {INPUT_FILE}")
print(f"Archivo de salida: {OUTPUT_FILE}")
print(f"Hoja procesada: {ws.title}")
print(f"Filas procesadas: {FIRST_DATA_ROW}-{LAST_DATA_ROW}")
print(f"Total de casos: {total_rows}")
print()

for item in summary_rows:
    print(
        f"{item['category']}: "
        f"{item['count']} casos "
        f"({item['percentage']:.2f}%)"
    )

if problem_rows:
    print()
    print("ADVERTENCIA: se han encontrado filas con valores problemáticos.")
    print("Primeras filas problemáticas:")
    for item in problem_rows[:20]:
        print(item)

print()
print(f"Resumen guardado en: {OUTPUT_SUMMARY_CSV}")