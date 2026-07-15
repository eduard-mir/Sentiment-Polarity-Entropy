"""
Calcula la accuracy entre Claude_label y la Gold human label, de forma
general y desglosada por entropy_band (low / mid / high).

Uso en PyCharm:
1. Ajusta INPUT_FILE con la ruta a tu Excel.
2. Revisa que CLAUDE_COL, GOLD_COL y BAND_COL coincidan EXACTAMENTE con
   los nombres de columna de tu archivo (ver sección CONFIGURACIÓN).
3. Ejecuta el script. Se generará un nuevo Excel (OUTPUT_FILE) con la
   tabla resumen.

Requiere: pandas, openpyxl  (pip install pandas openpyxl)
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ----------------------- CONFIGURACIÓN -----------------------
INPUT_FILE = "LLM_prueba_anotation_Claude.xlsx"   # archivo de entrada
OUTPUT_FILE = "accuracy_summary.xlsx"             # archivo de salida

CLAUDE_COL = "Claude_label"       # columna C
GOLD_COL = "Gold human label"     # columna D
BAND_COL = "entropy_band"         # columna E (ajusta si es otra)

BAND_ORDER = ["low", "mid", "high"]   # orden deseado en la tabla
# ---------------------------------------------------------------


def find_column(df, expected_name):
    """Busca una columna por nombre exacto; si no existe, intenta un
    match insensible a mayúsculas/espacios para tolerar pequeñas
    diferencias de encabezado."""
    if expected_name in df.columns:
        return expected_name
    normalized = {str(c).strip().lower(): c for c in df.columns}
    key = expected_name.strip().lower()
    if key in normalized:
        return normalized[key]
    return None


def main():
    if not Path(INPUT_FILE).exists():
        sys.exit(f"No se encontró el archivo de entrada: {INPUT_FILE}")

    df = pd.read_excel(INPUT_FILE)

    resolved = {}
    missing = []
    for label, expected in [("Claude_label", CLAUDE_COL),
                             ("Gold label", GOLD_COL),
                             ("entropy_band", BAND_COL)]:
        found = find_column(df, expected)
        if found is None:
            missing.append(expected)
        else:
            resolved[label] = found

    if missing:
        sys.exit(
            "No se encontraron estas columnas en el excel: "
            f"{missing}\nColumnas disponibles: {list(df.columns)}\n"
            "Ajusta CLAUDE_COL / GOLD_COL / BAND_COL en la sección "
            "CONFIGURACIÓN para que coincidan con tu archivo."
        )

    claude_col = resolved["Claude_label"]
    gold_col = resolved["Gold label"]
    band_col = resolved["entropy_band"]

    df = df.dropna(subset=[claude_col, gold_col]).copy()

    claude_norm = df[claude_col].astype(str).str.strip().str.upper()
    gold_norm = df[gold_col].astype(str).str.strip().str.upper()
    df["_correct"] = claude_norm == gold_norm

    overall_n = len(df)
    overall_correct = int(df["_correct"].sum())
    overall_acc = overall_correct / overall_n if overall_n else float("nan")

    rows = [{
        "Grupo": "General (todas las filas)",
        "N": overall_n,
        "Aciertos": overall_correct,
        "Accuracy": overall_acc,
    }]

    band_norm = df[band_col].astype(str).str.strip().str.lower()
    present = [b for b in BAND_ORDER if b in set(band_norm)]
    extra = sorted(b for b in set(band_norm) if b not in BAND_ORDER and b != "nan")
    ordered_bands = present + extra

    for band in ordered_bands:
        sub = df[band_norm == band]
        n = len(sub)
        correct = int(sub["_correct"].sum())
        acc = correct / n if n else float("nan")
        rows.append({
            "Grupo": f"entropy_band = {band}",
            "N": n,
            "Aciertos": correct,
            "Accuracy": acc,
        })

    summary = pd.DataFrame(rows)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Resumen")

    _format_output(OUTPUT_FILE, n_rows=len(summary))

    print(summary.to_string(index=False))
    print(f"\nResumen guardado en: {Path(OUTPUT_FILE).resolve()}")


def _format_output(path, n_rows):
    """Aplica formato básico (fuente, negrita en cabecera, % en Accuracy)."""
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="4472C4")
    body_font = Font(name="Arial", size=11)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
        for cell in row:
            cell.font = body_font
            if cell.column_letter == "D":  # columna Accuracy
                cell.number_format = "0.00%"
            cell.alignment = Alignment(horizontal="center" if cell.column_letter != "A" else "left")

    widths = {"A": 30, "B": 10, "C": 12, "D": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(path)


if __name__ == "__main__":
    main()