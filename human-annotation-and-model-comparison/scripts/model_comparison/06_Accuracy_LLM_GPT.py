from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# =========================
# CONFIGURACIÓN
# =========================

INPUT_FILE = Path("LLM_prueba_anotation_CHATGPT.xlsx")
OUTPUT_FILE = Path("chatgpt_vs_gold_accuracy_summary.xlsx")

SHEET_NAME = None  # None = usa la hoja activa

FIRST_DATA_ROW = 2
LAST_DATA_ROW = 960

CHATGPT_LABEL_COL = "C"
GOLD_LABEL_COL = "D"
ENTROPY_BAND_COL = "E"  # Cambia esta letra si entropy_band está en otra columna

VALID_LABELS = {"NEG", "NEU", "POS"}
ENTROPY_BANDS = ["low", "mid", "high"]


# =========================
# FUNCIONES
# =========================

def clean_label(value):
    if value is None:
        return None

    value = str(value).strip().upper()

    if value in VALID_LABELS:
        return value

    return None


def clean_entropy_band(value):
    if value is None:
        return None

    value = str(value).strip().lower()

    if value in ENTROPY_BANDS:
        return value

    return None


def percentage(correct, total):
    if total == 0:
        return 0.0
    return (correct / total) * 100


def style_header(row):
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")


def apply_basic_style(ws):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 3, 30)


# =========================
# LEER EXCEL
# =========================

wb_input = load_workbook(INPUT_FILE, data_only=True)

if SHEET_NAME is None:
    ws_input = wb_input.active
else:
    ws_input = wb_input[SHEET_NAME]


# =========================
# INICIALIZAR RESULTADOS
# =========================

overall = {
    "total_cases": 0,
    "correct_cases": 0,
    "incorrect_cases": 0,
}

by_entropy_band = {
    band: {
        "total_cases": 0,
        "correct_cases": 0,
        "incorrect_cases": 0,
    }
    for band in ENTROPY_BANDS
}

invalid_rows = []


# =========================
# CALCULAR ACCURACY
# =========================

for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):

    claude_label = clean_label(ws_input[f"{CHATGPT_LABEL_COL}{row}"].value)
    gold_label = clean_label(ws_input[f"{GOLD_LABEL_COL}{row}"].value)
    entropy_band = clean_entropy_band(ws_input[f"{ENTROPY_BAND_COL}{row}"].value)

    if claude_label is None or gold_label is None:
        invalid_rows.append({
            "excel_row": row,
            "chatgpt_label_raw": ws_input[f"{CHATGPT_LABEL_COL}{row}"].value,
            "gold_label_raw": ws_input[f"{GOLD_LABEL_COL}{row}"].value,
            "entropy_band_raw": ws_input[f"{ENTROPY_BAND_COL}{row}"].value,
            "reason": "missing_or_invalid_label",
        })
        continue

    is_correct = chatgpt_label == gold_label

    overall["total_cases"] += 1

    if is_correct:
        overall["correct_cases"] += 1
    else:
        overall["incorrect_cases"] += 1

    if entropy_band is None:
        invalid_rows.append({
            "excel_row": row,
            "chatgpt_label_raw": ws_input[f"{CHATGPT_LABEL_COL}{row}"].value,
            "gold_label_raw": ws_input[f"{GOLD_LABEL_COL}{row}"].value,
            "entropy_band_raw": ws_input[f"{ENTROPY_BAND_COL}{row}"].value,
            "reason": "missing_or_invalid_entropy_band",
        })
        continue

    by_entropy_band[entropy_band]["total_cases"] += 1

    if is_correct:
        by_entropy_band[entropy_band]["correct_cases"] += 1
    else:
        by_entropy_band[entropy_band]["incorrect_cases"] += 1


# =========================
# CREAR EXCEL DE SALIDA
# =========================

wb_output = Workbook()
ws_summary = wb_output.active
ws_summary.title = "accuracy_summary"

headers = [
    "set",
    "entropy_band",
    "total_cases",
    "correct_cases",
    "incorrect_cases",
    "accuracy_percent",
]

ws_summary.append(headers)
style_header(ws_summary[1])

ws_summary.append([
    "overall",
    "all",
    overall["total_cases"],
    overall["correct_cases"],
    overall["incorrect_cases"],
    percentage(overall["correct_cases"], overall["total_cases"]),
])

for band in ENTROPY_BANDS:
    total_cases = by_entropy_band[band]["total_cases"]
    correct_cases = by_entropy_band[band]["correct_cases"]
    incorrect_cases = by_entropy_band[band]["incorrect_cases"]

    ws_summary.append([
        "by_entropy_band",
        band,
        total_cases,
        correct_cases,
        incorrect_cases,
        percentage(correct_cases, total_cases),
    ])

for row in range(2, ws_summary.max_row + 1):
    ws_summary[f"F{row}"].number_format = "0.00"

apply_basic_style(ws_summary)


# =========================
# HOJA DE FILAS INVÁLIDAS
# =========================

ws_invalid = wb_output.create_sheet("invalid_rows")

ws_invalid.append([
    "excel_row",
    "chatgpt_label_raw",
    "gold_label_raw",
    "entropy_band_raw",
    "reason",
])

style_header(ws_invalid[1])

for item in invalid_rows:
    ws_invalid.append([
        item["excel_row"],
        item["chatgpt_label_raw"],
        item["gold_label_raw"],
        item["entropy_band_raw"],
        item["reason"],
    ])

apply_basic_style(ws_invalid)


# =========================
# GUARDAR EXCEL
# =========================

wb_output.save(OUTPUT_FILE)


# =========================
# RESULTADOS EN CONSOLA
# =========================

print("CHATGPT_LABEL VS GOLD HUMAN LABEL")
print("=================================")
print(f"Archivo analizado: {INPUT_FILE}")
print(f"Hoja analizada: {ws_input.title}")
print(f"Filas analizadas: {FIRST_DATA_ROW}-{LAST_DATA_ROW}")
print()

overall_accuracy = percentage(
    overall["correct_cases"],
    overall["total_cases"]
)

print("ACCURACY GENERAL")
print(f"Total cases: {overall['total_cases']}")
print(f"Correct cases: {overall['correct_cases']}")
print(f"Incorrect cases: {overall['incorrect_cases']}")
print(f"Accuracy: {overall_accuracy:.2f}%")
print()

print("ACCURACY BY ENTROPY_BAND")
for band in ENTROPY_BANDS:
    total_cases = by_entropy_band[band]["total_cases"]
    correct_cases = by_entropy_band[band]["correct_cases"]
    incorrect_cases = by_entropy_band[band]["incorrect_cases"]
    accuracy_percent = percentage(correct_cases, total_cases)

    print(
        f"{band}: "
        f"total_cases={total_cases}, "
        f"correct_cases={correct_cases}, "
        f"incorrect_cases={incorrect_cases}, "
        f"accuracy={accuracy_percent:.2f}%"
    )

if invalid_rows:
    print()
    print(f"ADVERTENCIA: {len(invalid_rows)} filas con valores inválidos o vacíos.")
    print("Primeros casos:")
    for item in invalid_rows[:20]:
        print(item)

print()
print(f"Archivo creado: {OUTPUT_FILE}")