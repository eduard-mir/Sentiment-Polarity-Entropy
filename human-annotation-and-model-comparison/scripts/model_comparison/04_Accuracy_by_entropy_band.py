from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# =========================
# CONFIGURACIÓN
# =========================

INPUT_FILE = Path(
    "sample_1000_sentences_for_manual_annotation_randomized_auto_label_gold_entropy_band_Comparsion.xlsx"
)

OUTPUT_FILE = Path(
    "cardiff_accuracy_by_entropy_band.xlsx"
)

SHEET_NAME = None  # None = usa la hoja activa

FIRST_DATA_ROW = 2
LAST_DATA_ROW = 960

ACCURACY_COL = "H"       # Accuracy_Cardiff
ENTROPY_BAND_COL = "J"   # entropy_band

ENTROPY_BANDS = ["low", "mid", "high"]


# =========================
# FUNCIONES
# =========================

def clean_accuracy(value):
    """
    Normaliza valores TRUE / FALSE de la columna Accuracy_Cardiff.
    Acepta booleanos reales de Excel o texto TRUE/FALSE.
    """
    if value is None:
        return None

    if isinstance(value, bool):
        return value

    value = str(value).strip().upper()

    if value == "TRUE":
        return True

    if value == "FALSE":
        return False

    return None


def clean_entropy_band(value):
    """
    Normaliza entropy_band a low / mid / high.
    """
    if value is None:
        return None

    value = str(value).strip().lower()

    if value in ENTROPY_BANDS:
        return value

    return None


def percentage(correct, total):
    if total == 0:
        return 0.0
    return (correct / total) * 100


def style_header(row):
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")


def apply_basic_style(ws):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 3, 28)


# =========================
# LEER EXCEL
# =========================

wb_input = load_workbook(INPUT_FILE, data_only=True)

if SHEET_NAME is None:
    ws_input = wb_input.active
else:
    ws_input = wb_input[SHEET_NAME]


# =========================
# CALCULAR ACCURACY POR ENTROPY_BAND
# =========================

results = {
    band: {
        "total_cases": 0,
        "correct_cases": 0,
        "incorrect_cases": 0,
    }
    for band in ENTROPY_BANDS
}

invalid_rows = []

for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):

    accuracy_value = clean_accuracy(ws_input[f"{ACCURACY_COL}{row}"].value)
    entropy_band = clean_entropy_band(ws_input[f"{ENTROPY_BAND_COL}{row}"].value)

    if accuracy_value is None or entropy_band is None:
        invalid_rows.append({
            "excel_row": row,
            "accuracy_raw": ws_input[f"{ACCURACY_COL}{row}"].value,
            "entropy_band_raw": ws_input[f"{ENTROPY_BAND_COL}{row}"].value,
        })
        continue

    results[entropy_band]["total_cases"] += 1

    if accuracy_value is True:
        results[entropy_band]["correct_cases"] += 1
    else:
        results[entropy_band]["incorrect_cases"] += 1


# =========================
# CREAR EXCEL DE SALIDA
# =========================

wb_output = Workbook()
ws_output = wb_output.active
ws_output.title = "accuracy_by_entropy_band"

headers = [
    "entropy_band",
    "total_cases",
    "correct_cases",
    "incorrect_cases",
    "accuracy_percent %",
]

ws_output.append(headers)
style_header(ws_output[1])

for band in ENTROPY_BANDS:
    total_cases = results[band]["total_cases"]
    correct_cases = results[band]["correct_cases"]
    incorrect_cases = results[band]["incorrect_cases"]
    accuracy_percent = percentage(correct_cases, total_cases)

    ws_output.append([
        band,
        total_cases,
        correct_cases,
        incorrect_cases,
        accuracy_percent,
    ])

for row in range(2, ws_output.max_row + 1):
    ws_output[f"E{row}"].number_format = "0.00"

apply_basic_style(ws_output)


# =========================
# HOJA CON FILAS INVÁLIDAS
# =========================

ws_invalid = wb_output.create_sheet("invalid_rows")

ws_invalid.append([
    "excel_row",
    "accuracy_raw",
    "entropy_band_raw",
])

style_header(ws_invalid[1])

for item in invalid_rows:
    ws_invalid.append([
        item["excel_row"],
        item["accuracy_raw"],
        item["entropy_band_raw"],
    ])

apply_basic_style(ws_invalid)


# =========================
# GUARDAR EXCEL
# =========================

wb_output.save(OUTPUT_FILE)


# =========================
# RESULTADOS EN CONSOLA
# =========================

print("CARDIFF ACCURACY BY ENTROPY_BAND")
print("================================")
print(f"Archivo analizado: {INPUT_FILE}")
print(f"Hoja analizada: {ws_input.title}")
print(f"Filas analizadas: {FIRST_DATA_ROW}-{LAST_DATA_ROW}")
print()

for band in ENTROPY_BANDS:
    total_cases = results[band]["total_cases"]
    correct_cases = results[band]["correct_cases"]
    incorrect_cases = results[band]["incorrect_cases"]
    accuracy_percent = percentage(correct_cases, total_cases)

    print(
        f"{band}: "
        f"total_cases={total_cases}, "
        f"correct_cases={correct_cases}, "
        f"incorrect_cases={incorrect_cases}, "
        f"accuracy={accuracy_percent:.2f}%"
    )

if invalid_rows:
    print()
    print(f"ADVERTENCIA: {len(invalid_rows)} filas con valores inválidos o vacíos.")
    print("Primeros casos:")
    for item in invalid_rows[:20]:
        print(item)

print()
print(f"Archivo creado: {OUTPUT_FILE}")