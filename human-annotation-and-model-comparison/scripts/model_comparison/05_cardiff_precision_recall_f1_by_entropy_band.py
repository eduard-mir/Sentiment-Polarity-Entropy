from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# =========================
# CONFIGURACIÓN
# =========================

INPUT_FILE = Path(
    "sample_1000_sentences_for_manual_annotation_randomized_auto_label_gold_entropy_band_Comparsion.xlsx"
)

OUTPUT_FILE = Path(
    "cardiff_precision_recall_f1_by_entropy_band.xlsx"
)

SHEET_NAME = None  # None = usa la hoja activa

FIRST_DATA_ROW = 2
LAST_DATA_ROW = 960

CARDIFF_LABEL_COL = "G"
GOLD_LABEL_COL = "I"
ENTROPY_BAND_COL = "J"

LABELS = ["POS", "NEU", "NEG"]
ENTROPY_BANDS = ["low", "mid", "high"]


# =========================
# FUNCIONES
# =========================

def clean_label(value):
    if value is None:
        return None

    value = str(value).strip().upper()

    if value in LABELS:
        return value

    return None


def clean_entropy_band(value):
    if value is None:
        return None

    value = str(value).strip().lower()

    if value in ENTROPY_BANDS:
        return value

    return None


def safe_divide(numerator, denominator):
    if denominator == 0:
        return 0.0
    return numerator / denominator


def compute_metrics_for_label(gold_labels, predicted_labels, target_label):
    tp = 0
    fp = 0
    fn = 0
    tn = 0

    for gold, pred in zip(gold_labels, predicted_labels):

        if gold == target_label and pred == target_label:
            tp += 1

        elif gold != target_label and pred == target_label:
            fp += 1

        elif gold == target_label and pred != target_label:
            fn += 1

        else:
            tn += 1

    precision = safe_divide(tp, tp + fp)
    recall = safe_divide(tp, tp + fn)

    if precision + recall == 0:
        f1 = 0.0
    else:
        f1 = 2 * precision * recall / (precision + recall)

    return {
        "gold_support": tp + fn,
        "auto_predicted": tp + fp,
        "true_positives": tp,
        "false_positives": fp,
        "false_negatives": fn,
        "true_negatives": tn,
        "precision": precision,
        "recall": recall,
        "f1": f1,
    }


def style_header(row):
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")


def apply_basic_style(ws):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 3, 30)


# =========================
# LEER EXCEL
# =========================

wb_input = load_workbook(INPUT_FILE, data_only=True)

if SHEET_NAME is None:
    ws_input = wb_input.active
else:
    ws_input = wb_input[SHEET_NAME]


# =========================
# AGRUPAR DATOS POR ENTROPY_BAND
# =========================

data_by_band = {
    band: {
        "gold_labels": [],
        "cardiff_labels": [],
    }
    for band in ENTROPY_BANDS
}

invalid_rows = []

for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):

    cardiff_label = clean_label(ws_input[f"{CARDIFF_LABEL_COL}{row}"].value)
    gold_label = clean_label(ws_input[f"{GOLD_LABEL_COL}{row}"].value)
    entropy_band = clean_entropy_band(ws_input[f"{ENTROPY_BAND_COL}{row}"].value)

    if cardiff_label is None or gold_label is None or entropy_band is None:
        invalid_rows.append({
            "excel_row": row,
            "cardiff_label_raw": ws_input[f"{CARDIFF_LABEL_COL}{row}"].value,
            "gold_label_raw": ws_input[f"{GOLD_LABEL_COL}{row}"].value,
            "entropy_band_raw": ws_input[f"{ENTROPY_BAND_COL}{row}"].value,
        })
        continue

    data_by_band[entropy_band]["gold_labels"].append(gold_label)
    data_by_band[entropy_band]["cardiff_labels"].append(cardiff_label)


# =========================
# CALCULAR MÉTRICAS
# =========================

result_rows = []

for band in ENTROPY_BANDS:

    gold_labels = data_by_band[band]["gold_labels"]
    cardiff_labels = data_by_band[band]["cardiff_labels"]

    for label in LABELS:

        metrics = compute_metrics_for_label(
            gold_labels=gold_labels,
            predicted_labels=cardiff_labels,
            target_label=label
        )

        result_rows.append({
            "entropy_band": band,
            "label": label,
            "gold_support": metrics["gold_support"],
            "auto_predicted": metrics["auto_predicted"],
            "true_positives": metrics["true_positives"],
            "false_positives": metrics["false_positives"],
            "false_negatives": metrics["false_negatives"],
            "true_negatives": metrics["true_negatives"],
            "precision": metrics["precision"],
            "recall": metrics["recall"],
            "f1": metrics["f1"],
        })


# =========================
# CREAR EXCEL DE SALIDA
# =========================

wb_output = Workbook()
ws_output = wb_output.active
ws_output.title = "metrics_by_entropy_band"

headers = [
    "entropy_band",
    "label",
    "gold_support",
    "auto_predicted",
    "true_positives",
    "false_positives",
    "false_negatives",
    "true_negatives",
    "precision",
    "recall",
    "f1",
]

ws_output.append(headers)
style_header(ws_output[1])

for item in result_rows:
    ws_output.append([
        item["entropy_band"],
        item["label"],
        item["gold_support"],
        item["auto_predicted"],
        item["true_positives"],
        item["false_positives"],
        item["false_negatives"],
        item["true_negatives"],
        item["precision"],
        item["recall"],
        item["f1"],
    ])

for row in range(2, ws_output.max_row + 1):
    ws_output[f"I{row}"].number_format = "0.0000"
    ws_output[f"J{row}"].number_format = "0.0000"
    ws_output[f"K{row}"].number_format = "0.0000"

apply_basic_style(ws_output)


# =========================
# HOJA DE FILAS INVÁLIDAS
# =========================

ws_invalid = wb_output.create_sheet("invalid_rows")

ws_invalid.append([
    "excel_row",
    "cardiff_label_raw",
    "gold_label_raw",
    "entropy_band_raw",
])

style_header(ws_invalid[1])

for item in invalid_rows:
    ws_invalid.append([
        item["excel_row"],
        item["cardiff_label_raw"],
        item["gold_label_raw"],
        item["entropy_band_raw"],
    ])

apply_basic_style(ws_invalid)


# =========================
# GUARDAR EXCEL
# =========================

wb_output.save(OUTPUT_FILE)


# =========================
# RESULTADOS EN CONSOLA
# =========================

print("CARDIFF PRECISION, RECALL Y F1 BY ENTROPY_BAND")
print("==============================================")
print(f"Archivo analizado: {INPUT_FILE}")
print(f"Hoja analizada: {ws_input.title}")
print(f"Filas analizadas: {FIRST_DATA_ROW}-{LAST_DATA_ROW}")
print()

for band in ENTROPY_BANDS:
    total_cases = len(data_by_band[band]["gold_labels"])
    print(f"{band}: {total_cases} casos")

print()

for item in result_rows:
    print(
        f"{item['entropy_band']} - {item['label']}: "
        f"precision={item['precision']:.4f}, "
        f"recall={item['recall']:.4f}, "
        f"f1={item['f1']:.4f}"
    )

if invalid_rows:
    print()
    print(f"ADVERTENCIA: {len(invalid_rows)} filas con valores inválidos o vacíos.")
    print("Primeros casos:")
    for item in invalid_rows[:20]:
        print(item)

print()
print(f"Archivo creado: {OUTPUT_FILE}")