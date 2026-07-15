from pathlib import Path
from openpyxl import load_workbook


# =========================
# CONFIGURACIÓN
# =========================

INPUT_FILE = Path(
    "sample_1000_sentences_for_manual_annotation_randomized_auto_label_gold_entropy_band_Comparsion.xlsx"
)

OUTPUT_FILE = Path(
    "sample_1000_sentences_for_manual_annotation_randomized_auto_label_gold_entropy_band_Comparsion_cardiff_accuracy.xlsx"
)

SHEET_NAME = None  # None = usa la hoja activa

FIRST_DATA_ROW = 2
LAST_DATA_ROW = 960

CARDIFF_LABEL_COL = "G"
GOLD_LABEL_COL = "H"
OUTPUT_MATCH_COL = "J"

VALID_LABELS = {"NEG", "NEU", "POS"}


# =========================
# FUNCIONES
# =========================

def clean_label(value):
    """
    Normaliza etiquetas POS / NEU / NEG.
    """
    if value is None:
        return None

    value = str(value).strip().upper()

    if value in VALID_LABELS:
        return value

    return None


def percentage(count, total):
    return (count / total) * 100 if total > 0 else 0


# =========================
# LEER EXCEL
# =========================

wb = load_workbook(INPUT_FILE)

if SHEET_NAME is None:
    ws = wb.active
else:
    ws = wb[SHEET_NAME]


# =========================
# COMPARAR CARDIFF_LABEL VS GOLD HUMAN LABEL
# =========================

ws[f"{OUTPUT_MATCH_COL}1"] = "Cardiff_vs_Gold_Match"

true_count = 0
false_count = 0
invalid_or_missing_count = 0

for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):

    cardiff_label = clean_label(ws[f"{CARDIFF_LABEL_COL}{row}"].value)
    gold_label = clean_label(ws[f"{GOLD_LABEL_COL}{row}"].value)

    if cardiff_label is None or gold_label is None:
        ws[f"{OUTPUT_MATCH_COL}{row}"] = "INVALID_OR_MISSING"
        invalid_or_missing_count += 1
        continue

    if cardiff_label == gold_label:
        ws[f"{OUTPUT_MATCH_COL}{row}"] = "TRUE"
        true_count += 1
    else:
        ws[f"{OUTPUT_MATCH_COL}{row}"] = "FALSE"
        false_count += 1


valid_cases = true_count + false_count
accuracy = percentage(true_count, valid_cases)


# =========================
# GUARDAR NUEVO EXCEL
# =========================

wb.save(OUTPUT_FILE)


# =========================
# RESULTADOS EN CONSOLA
# =========================

print("CARDIFF_LABEL VS GOLD HUMAN LABEL")
print("=================================")
print(f"Archivo analizado: {INPUT_FILE}")
print(f"Archivo creado: {OUTPUT_FILE}")
print(f"Hoja analizada: {ws.title}")
print(f"Filas analizadas: {FIRST_DATA_ROW}-{LAST_DATA_ROW}")
print()
print(f"Casos válidos: {valid_cases}")
print(f"Casos TRUE: {true_count}")
print(f"Casos FALSE: {false_count}")
print(f"Casos inválidos o vacíos: {invalid_or_missing_count}")
print()
print(f"Accuracy Cardiff vs Gold: {accuracy:.2f}%")