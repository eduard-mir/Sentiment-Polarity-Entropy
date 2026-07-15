from pathlib import Path
import re
import unicodedata

import torch
from transformers import AutoTokenizer, AutoModelForSequenceClassification
from openpyxl import load_workbook
from tqdm import tqdm


# =========================
# CONFIGURACIÓN
# =========================

INPUT_FILE = Path(
    "sample_1000_sentences_for_manual_annotation_randomized_auto_label_gold_entropy_band_Comparsion.xlsx"
)

OUTPUT_FILE = Path(
    "sample_1000_sentences_for_manual_annotation_randomized_auto_label_gold_entropy_band_Comparsion_cardiff.xlsx"
)

MODEL_NAME = "cardiffnlp/twitter-xlm-roberta-base-sentiment"

SHEET_NAME = None  # None = usa la hoja activa

FIRST_DATA_ROW = 2
LAST_DATA_ROW = 960

TARGET_WORD_COL = "C"
SENTENCE_COL = "D"

# Insertamos nuevas columnas desde I.
# I = Cardiff_label
# J = Cardiff_NEG
# K = Cardiff_NEU
# L = Cardiff_POS
# M = Cardiff_context_window
# N = Cardiff_target_found
INSERT_AT_COLUMN_INDEX = 9
N_COLUMNS_TO_INSERT = 6

LEFT_WINDOW = 5
RIGHT_WINDOW = 5

BATCH_SIZE = 16
MAX_LENGTH = 128

USE_FULL_SENTENCE_IF_TARGET_NOT_FOUND = True


# =========================
# FUNCIONES DE NORMALIZACIÓN
# =========================

def normalize_text(value):
    """
    Normaliza texto para comparar palabras:
    - minúsculas
    - sin espacios externos
    - sin acentos
    """
    if value is None:
        return ""

    value = str(value).strip().lower()
    value = unicodedata.normalize("NFD", value)
    value = "".join(ch for ch in value if unicodedata.category(ch) != "Mn")
    return value


def candidate_forms(target):
    """
    Genera algunas variantes morfológicas simples para adjetivos españoles.
    Esto ayuda si la columna C contiene el lema, pero en la oración aparece
    una forma flexionada: bueno/buena/buenos/buenas, alto/alta/altos/altas, etc.
    """
    target = normalize_text(target)
    forms = {target}

    if target.endswith("o") and len(target) > 2:
        stem = target[:-1]
        forms.update({
            stem + "a",
            stem + "os",
            stem + "as",
            stem,        # posible forma apocopada: bueno -> buen
        })

    elif target.endswith("e") and len(target) > 2:
        forms.add(target + "s")

    elif target.endswith("z") and len(target) > 2:
        forms.add(target[:-1] + "ces")

    elif len(target) > 2:
        forms.add(target + "s")
        forms.add(target + "es")

    return forms


# =========================
# TOKENIZACIÓN LOCAL PARA VENTANA
# =========================

TOKEN_RE = re.compile(r"\w+|[^\w\s]", re.UNICODE)
WORD_RE = re.compile(r"^\w+$", re.UNICODE)


def tokenize_with_punctuation(text):
    if text is None:
        return []
    return TOKEN_RE.findall(str(text))


def is_word_token(token):
    return bool(WORD_RE.match(token))


def detokenize(tokens):
    """
    Reconstruye un fragmento evitando espacios antes de puntuación.
    """
    text = ""

    no_space_before = {".", ",", ";", ":", "!", "?", "%", ")", "]", "}", "»", "”"}
    no_space_after = {"¿", "¡", "(", "[", "{", "«", "“"}

    for token in tokens:
        if not text:
            text = token
        elif token in no_space_before:
            text += token
        elif text[-1] in no_space_after:
            text += token
        else:
            text += " " + token

    return text


def extract_context_window(sentence, target_word, left_window=5, right_window=5):
    """
    Busca la palabra objetivo en la oración y devuelve una ventana de:
    5 palabras a la izquierda + target + 5 palabras a la derecha.

    Si no encuentra el target, puede devolver la oración completa como fallback.
    """
    tokens = tokenize_with_punctuation(sentence)
    target_forms = candidate_forms(target_word)

    word_token_indices = [
        i for i, token in enumerate(tokens)
        if is_word_token(token)
    ]

    target_word_position = None
    target_token_index = None

    for word_position, token_index in enumerate(word_token_indices):
        token_norm = normalize_text(tokens[token_index])

        if token_norm in target_forms:
            target_word_position = word_position
            target_token_index = token_index
            break

    if target_word_position is None:
        if USE_FULL_SENTENCE_IF_TARGET_NOT_FOUND:
            return str(sentence), False
        else:
            return "", False

    left_word_position = max(0, target_word_position - left_window)
    right_word_position = min(
        len(word_token_indices) - 1,
        target_word_position + right_window
    )

    left_token_index = word_token_indices[left_word_position]
    right_token_index = word_token_indices[right_word_position]

    context_tokens = tokens[left_token_index:right_token_index + 1]
    context_text = detokenize(context_tokens)

    return context_text, True


# =========================
# MAPEO DE ETIQUETAS DEL MODELO
# =========================

def map_cardiff_label(label):
    """
    Cardiff suele devolver:
    negative / neutral / positive

    También dejamos contemplado el caso:
    LABEL_0 / LABEL_1 / LABEL_2
    """
    label = str(label).strip().lower()

    mapping = {
        "negative": "NEG",
        "neutral": "NEU",
        "positive": "POS",
        "neg": "NEG",
        "neu": "NEU",
        "pos": "POS",
        "label_0": "NEG",
        "label_1": "NEU",
        "label_2": "POS",
    }

    if label not in mapping:
        raise ValueError(f"No sé mapear esta etiqueta del modelo: {label}")

    return mapping[label]


# =========================
# CARGAR MODELO
# =========================

print("Cargando modelo...")
print(f"Modelo: {MODEL_NAME}")

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Dispositivo: {device}")

tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME, use_fast=False)
model = AutoModelForSequenceClassification.from_pretrained(MODEL_NAME)
model.to(device)
model.eval()

id2label = model.config.id2label
print(f"id2label del modelo: {id2label}")


# =========================
# LEER EXCEL
# =========================

wb = load_workbook(INPUT_FILE)
ws = wb.active if SHEET_NAME is None else wb[SHEET_NAME]

print()
print(f"Archivo de entrada: {INPUT_FILE}")
print(f"Hoja: {ws.title}")
print(f"Filas procesadas: {FIRST_DATA_ROW}-{LAST_DATA_ROW}")


# =========================
# EXTRAER FRAGMENTOS
# =========================

rows_to_process = []
contexts = []
target_found_flags = []

for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
    target_word = ws[f"{TARGET_WORD_COL}{row}"].value
    sentence = ws[f"{SENTENCE_COL}{row}"].value

    context_window, target_found = extract_context_window(
        sentence=sentence,
        target_word=target_word,
        left_window=LEFT_WINDOW,
        right_window=RIGHT_WINDOW
    )

    rows_to_process.append(row)
    contexts.append(context_window)
    target_found_flags.append(target_found)


# =========================
# APLICAR ANALIZADOR CARDIFF
# =========================

all_results = []

print()
print("Aplicando analizador de sentimiento...")

with torch.no_grad():
    for start in tqdm(range(0, len(contexts), BATCH_SIZE)):
        batch_contexts = contexts[start:start + BATCH_SIZE]

        encoded = tokenizer(
            batch_contexts,
            padding=True,
            truncation=True,
            max_length=MAX_LENGTH,
            return_tensors="pt"
        )

        encoded = {
            key: value.to(device)
            for key, value in encoded.items()
        }

        outputs = model(**encoded)
        probabilities = torch.softmax(outputs.logits, dim=-1)

        for probs in probabilities:
            score_by_label = {
                "NEG": 0.0,
                "NEU": 0.0,
                "POS": 0.0,
            }

            for class_id, score in enumerate(probs.tolist()):
                raw_label = id2label[class_id]
                mapped_label = map_cardiff_label(raw_label)
                score_by_label[mapped_label] = float(score)

            predicted_label = max(
                score_by_label,
                key=lambda label: score_by_label[label]
            )

            all_results.append({
                "label": predicted_label,
                "NEG": score_by_label["NEG"],
                "NEU": score_by_label["NEU"],
                "POS": score_by_label["POS"],
            })


# =========================
# INSERTAR COLUMNAS Y GUARDAR RESULTADOS
# =========================

ws.insert_cols(INSERT_AT_COLUMN_INDEX, amount=N_COLUMNS_TO_INSERT)

ws.cell(row=1, column=INSERT_AT_COLUMN_INDEX).value = "Cardiff_label"
ws.cell(row=1, column=INSERT_AT_COLUMN_INDEX + 1).value = "Cardiff_NEG"
ws.cell(row=1, column=INSERT_AT_COLUMN_INDEX + 2).value = "Cardiff_NEU"
ws.cell(row=1, column=INSERT_AT_COLUMN_INDEX + 3).value = "Cardiff_POS"
ws.cell(row=1, column=INSERT_AT_COLUMN_INDEX + 4).value = "Cardiff_context_window"
ws.cell(row=1, column=INSERT_AT_COLUMN_INDEX + 5).value = "Cardiff_target_found"

for row, result, context, target_found in zip(
    rows_to_process,
    all_results,
    contexts,
    target_found_flags
):
    ws.cell(row=row, column=INSERT_AT_COLUMN_INDEX).value = result["label"]
    ws.cell(row=row, column=INSERT_AT_COLUMN_INDEX + 1).value = result["NEG"]
    ws.cell(row=row, column=INSERT_AT_COLUMN_INDEX + 2).value = result["NEU"]
    ws.cell(row=row, column=INSERT_AT_COLUMN_INDEX + 3).value = result["POS"]
    ws.cell(row=row, column=INSERT_AT_COLUMN_INDEX + 4).value = context
    ws.cell(row=row, column=INSERT_AT_COLUMN_INDEX + 5).value = target_found

    ws.cell(row=row, column=INSERT_AT_COLUMN_INDEX + 1).number_format = "0.0000"
    ws.cell(row=row, column=INSERT_AT_COLUMN_INDEX + 2).number_format = "0.0000"
    ws.cell(row=row, column=INSERT_AT_COLUMN_INDEX + 3).number_format = "0.0000"


# =========================
# RESUMEN
# =========================

total_cases = len(rows_to_process)
target_found_count = sum(1 for flag in target_found_flags if flag)
target_not_found_count = total_cases - target_found_count

label_counts = {
    "NEG": sum(1 for item in all_results if item["label"] == "NEG"),
    "NEU": sum(1 for item in all_results if item["label"] == "NEU"),
    "POS": sum(1 for item in all_results if item["label"] == "POS"),
}

wb.save(OUTPUT_FILE)

print()
print("ANÁLISIS COMPLETADO")
print("===================")
print(f"Archivo de salida: {OUTPUT_FILE}")
print(f"Total de casos procesados: {total_cases}")
print(f"Target encontrado: {target_found_count}")
print(f"Target no encontrado: {target_not_found_count}")
print()
print("Distribución Cardiff_label:")
print(f"NEG: {label_counts['NEG']}")
print(f"NEU: {label_counts['NEU']}")
print(f"POS: {label_counts['POS']}")