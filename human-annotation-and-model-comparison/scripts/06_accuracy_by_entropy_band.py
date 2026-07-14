from pathlib import Path
from openpyxl import load_workbook
import csv


# =========================
# CONFIGURACIÓN
# =========================

INPUT_FILE = Path("sample_1000_sentences_for_manual_annotation_randomized_auto_label_gold_entropy_band.xlsx")
OUTPUT_CSV = Path("accuracy_by_entropy_band.csv")

SHEET_NAME = None  # None = usa la hoja activa

FIRST_DATA_ROW = 2
LAST_DATA_ROW = 960

ACCURACY_COL = "P"
ENTROPY_BAND_COL = "Q"

VALID_BANDS = ["low", "mid", "high"]


# =========================
# FUNCIONES
# =========================

def normalize_entropy_band(value):
    if value is None:
        return None

    value = str(value).strip().lower()

    if value in VALID_BANDS:
        return value

    return None


def normalize_accuracy(value):
    """
    Convierte TRUE / FALSE a booleano.
    """

    if value is None:
        return None

    if isinstance(value, bool):
        return value

    value = str(value).strip().upper()

    if value == "TRUE":
        return True

    if value == "FALSE":
        return False

    return None


def percentage(count, total):
    return (count / total) * 100 if total > 0 else 0


# =========================
# LECTURA DEL EXCEL
# =========================

wb = load_workbook(INPUT_FILE, data_only=True)

if SHEET_NAME is None:
    ws = wb.active
else:
    ws = wb[SHEET_NAME]


# =========================
# CÁLCULO DE ACCURACY POR BANDA
# =========================

results = {
    band: {
        "total_cases": 0,
        "correct_cases": 0,
        "incorrect_cases": 0,
    }
    for band in VALID_BANDS
}

invalid_rows = []

for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):

    accuracy_value = normalize_accuracy(ws[f"{ACCURACY_COL}{row}"].value)
    entropy_band = normalize_entropy_band(ws[f"{ENTROPY_BAND_COL}{row}"].value)

    if accuracy_value is None or entropy_band is None:
        invalid_rows.append({
            "excel_row": row,
            "accuracy_value": ws[f"{ACCURACY_COL}{row}"].value,
            "entropy_band": ws[f"{ENTROPY_BAND_COL}{row}"].value,
        })
        continue

    results[entropy_band]["total_cases"] += 1

    if accuracy_value is True:
        results[entropy_band]["correct_cases"] += 1
    else:
        results[entropy_band]["incorrect_cases"] += 1


# =========================
# PREPARAR RESUMEN
# =========================

summary_rows = []

for band in VALID_BANDS:
    total = results[band]["total_cases"]
    correct = results[band]["correct_cases"]
    incorrect = results[band]["incorrect_cases"]
    accuracy_percent = percentage(correct, total)

    summary_rows.append({
        "entropy_band": band,
        "total_cases": total,
        "correct_cases": correct,
        "incorrect_cases": incorrect,
        "accuracy_percent": accuracy_percent,
    })


# =========================
# IMPRIMIR RESULTADOS
# =========================

print("ACCURACY POR ENTROPY_BAND")
print("=========================")
print(f"Archivo analizado: {INPUT_FILE}")
print(f"Hoja analizada: {ws.title}")
print(f"Filas analizadas: {FIRST_DATA_ROW}-{LAST_DATA_ROW}")
print()

for item in summary_rows:
    print(
        f"{item['entropy_band']}: "
        f"{item['correct_cases']}/{item['total_cases']} correctos "
        f"({item['accuracy_percent']:.2f}%)"
    )

if invalid_rows:
    print()
    print("ADVERTENCIA: filas con valores no válidos o vacíos:")
    for item in invalid_rows[:20]:
        print(item)

    if len(invalid_rows) > 20:
        print(f"... y {len(invalid_rows) - 20} filas más.")


# =========================
# GUARDAR RESULTADOS
# =========================

with open(OUTPUT_CSV, mode="w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(
        f,
        fieldnames=[
            "entropy_band",
            "total_cases",
            "correct_cases",
            "incorrect_cases",
            "accuracy_percent",
        ],
        delimiter=";"
    )
    writer.writeheader()
    writer.writerows(summary_rows)


print()
print(f"Resumen guardado en: {OUTPUT_CSV}")