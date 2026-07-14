from pathlib import Path
from openpyxl import load_workbook
import csv


# =========================
# CONFIGURACIÓN
# =========================

INPUT_FILE = Path("sample_1000_sentences_for_manual_annotation_randomized_auto_label_gold.xlsx")

OUTPUT_SUMMARY_CSV = Path("auto_vs_gold_accuracy_summary.csv")
OUTPUT_CASES_CSV = Path("auto_vs_gold_accuracy_cases.csv")

SHEET_NAME = None  # None = usa la hoja activa

FIRST_DATA_ROW = 2
LAST_DATA_ROW = 960

AUTO_LABEL_COL = "I"
GOLD_LABEL_COL = "O"

VALID_LABELS = {"NEG", "NEU", "POS"}


# =========================
# FUNCIONES
# =========================

def clean_label(value):
    """
    Limpia etiquetas POS / NEU / NEG.
    """
    if value is None:
        return None

    value = str(value).strip().upper()

    if value in VALID_LABELS:
        return value

    return None


def percentage(count, total):
    return (count / total) * 100 if total > 0 else 0


# =========================
# LECTURA DEL EXCEL
# =========================

wb = load_workbook(INPUT_FILE, data_only=True)

if SHEET_NAME is None:
    ws = wb.active
else:
    ws = wb[SHEET_NAME]


# =========================
# CÁLCULO DE ACCURACY
# =========================

case_rows = []

correct_count = 0
incorrect_count = 0
invalid_count = 0

for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):

    auto_label = clean_label(ws[f"{AUTO_LABEL_COL}{row}"].value)
    gold_label = clean_label(ws[f"{GOLD_LABEL_COL}{row}"].value)

    if auto_label is None or gold_label is None:
        invalid_count += 1
        case_rows.append({
            "excel_row": row,
            "auto_label": auto_label,
            "gold_human_label": gold_label,
            "match": "INVALID_OR_MISSING",
        })
        continue

    is_correct = auto_label == gold_label

    if is_correct:
        correct_count += 1
        match_value = "TRUE"
    else:
        incorrect_count += 1
        match_value = "FALSE"

    case_rows.append({
        "excel_row": row,
        "auto_label": auto_label,
        "gold_human_label": gold_label,
        "match": match_value,
    })


valid_total = correct_count + incorrect_count
accuracy = percentage(correct_count, valid_total)


# =========================
# RESUMEN
# =========================

summary_rows = [
    {
        "metric": "total_rows_checked",
        "value": LAST_DATA_ROW - FIRST_DATA_ROW + 1,
    },
    {
        "metric": "valid_cases",
        "value": valid_total,
    },
    {
        "metric": "correct_cases",
        "value": correct_count,
    },
    {
        "metric": "incorrect_cases",
        "value": incorrect_count,
    },
    {
        "metric": "invalid_or_missing_cases",
        "value": invalid_count,
    },
    {
        "metric": "accuracy",
        "value": accuracy,
    },
]


# =========================
# IMPRIMIR RESULTADOS
# =========================

print("ACCURACY: AUTO_LABEL VS GOLD HUMAN LABEL")
print("========================================")
print(f"Archivo analizado: {INPUT_FILE}")
print(f"Hoja analizada: {ws.title}")
print(f"Filas analizadas: {FIRST_DATA_ROW}-{LAST_DATA_ROW}")
print()
print(f"Casos válidos: {valid_total}")
print(f"Casos correctos: {correct_count}")
print(f"Casos incorrectos: {incorrect_count}")
print(f"Casos inválidos o vacíos: {invalid_count}")
print(f"Accuracy: {accuracy:.2f}%")


# =========================
# GUARDAR RESUMEN
# =========================

with open(OUTPUT_SUMMARY_CSV, mode="w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(
        f,
        fieldnames=["metric", "value"],
        delimiter=";"
    )
    writer.writeheader()
    writer.writerows(summary_rows)


# =========================
# GUARDAR DETALLE POR CASO
# =========================

with open(OUTPUT_CASES_CSV, mode="w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(
        f,
        fieldnames=[
            "excel_row",
            "auto_label",
            "gold_human_label",
            "match",
        ],
        delimiter=";"
    )
    writer.writeheader()
    writer.writerows(case_rows)


print()
print(f"Resumen guardado en: {OUTPUT_SUMMARY_CSV}")
print(f"Detalle por caso guardado en: {OUTPUT_CASES_CSV}")