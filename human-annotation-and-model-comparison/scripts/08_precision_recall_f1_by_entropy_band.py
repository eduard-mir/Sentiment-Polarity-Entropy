from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# =========================
# CONFIGURACIÓN
# =========================

INPUT_FILE = Path("sample_1000_sentences_for_manual_annotation_randomized_auto_label_gold_entropy_band.xlsx")
OUTPUT_FILE = Path("auto_vs_gold_precision_recall_f1_by_entropy_band.xlsx")

SHEET_NAME = None  # None = usa la hoja activa

FIRST_DATA_ROW = 2
LAST_DATA_ROW = 960

AUTO_LABEL_COL = "I"
GOLD_LABEL_COL = "O"
ENTROPY_BAND_COL = "Q"

LABELS = ["NEG", "NEU", "POS"]
ENTROPY_BANDS = ["low", "mid", "high"]


# =========================
# FUNCIONES
# =========================

def clean_label(value):
    if value is None:
        return None

    value = str(value).strip().upper()

    if value in LABELS:
        return value

    return None


def clean_entropy_band(value):
    if value is None:
        return None

    value = str(value).strip().lower()

    if value in ENTROPY_BANDS:
        return value

    return None


def safe_divide(numerator, denominator):
    if denominator == 0:
        return 0.0
    return numerator / denominator


def compute_metrics_for_label(gold_labels, auto_labels, target_label):
    """
    Calcula TP, FP, FN, TN, precision, recall y F1 para una etiqueta.
    """

    tp = fp = fn = tn = 0

    for gold, auto in zip(gold_labels, auto_labels):

        if gold == target_label and auto == target_label:
            tp += 1
        elif gold != target_label and auto == target_label:
            fp += 1
        elif gold == target_label and auto != target_label:
            fn += 1
        else:
            tn += 1

    precision = safe_divide(tp, tp + fp)
    recall = safe_divide(tp, tp + fn)

    if precision + recall == 0:
        f1 = 0.0
    else:
        f1 = 2 * precision * recall / (precision + recall)

    return {
        "label": target_label,
        "gold_support": tp + fn,
        "auto_predicted": tp + fp,
        "true_positives": tp,
        "false_positives": fp,
        "false_negatives": fn,
        "true_negatives": tn,
        "precision": precision,
        "recall": recall,
        "f1": f1,
    }


def build_confusion_matrix(gold_labels, auto_labels):
    """
    Matriz de confusión:
    filas = Gold Human Label
    columnas = auto_label
    """

    matrix = {
        gold_label: {auto_label: 0 for auto_label in LABELS}
        for gold_label in LABELS
    }

    for gold, auto in zip(gold_labels, auto_labels):
        matrix[gold][auto] += 1

    return matrix


def style_header(row):
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")


def apply_basic_style(ws):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 3, 32)


# =========================
# LECTURA DEL EXCEL
# =========================

wb_input = load_workbook(INPUT_FILE, data_only=True)

if SHEET_NAME is None:
    ws_input = wb_input.active
else:
    ws_input = wb_input[SHEET_NAME]


# =========================
# CLASIFICAR FILAS POR ENTROPY_BAND
# =========================

data_by_band = {
    band: {
        "gold_labels": [],
        "auto_labels": [],
        "excel_rows": [],
    }
    for band in ENTROPY_BANDS
}

invalid_rows = []

for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):

    auto_label = clean_label(ws_input[f"{AUTO_LABEL_COL}{row}"].value)
    gold_label = clean_label(ws_input[f"{GOLD_LABEL_COL}{row}"].value)
    entropy_band = clean_entropy_band(ws_input[f"{ENTROPY_BAND_COL}{row}"].value)

    if auto_label is None or gold_label is None or entropy_band is None:
        invalid_rows.append({
            "excel_row": row,
            "auto_label_raw": ws_input[f"{AUTO_LABEL_COL}{row}"].value,
            "gold_label_raw": ws_input[f"{GOLD_LABEL_COL}{row}"].value,
            "entropy_band_raw": ws_input[f"{ENTROPY_BAND_COL}{row}"].value,
        })
        continue

    data_by_band[entropy_band]["gold_labels"].append(gold_label)
    data_by_band[entropy_band]["auto_labels"].append(auto_label)
    data_by_band[entropy_band]["excel_rows"].append(row)


# =========================
# CALCULAR MÉTRICAS POR BANDA
# =========================

all_metric_rows = []
summary_rows = []
confusion_matrices = {}

for band in ENTROPY_BANDS:

    gold_labels = data_by_band[band]["gold_labels"]
    auto_labels = data_by_band[band]["auto_labels"]

    total_cases = len(gold_labels)
    correct_cases = sum(
        1 for gold, auto in zip(gold_labels, auto_labels)
        if gold == auto
    )

    accuracy = safe_divide(correct_cases, total_cases)

    metrics = [
        compute_metrics_for_label(gold_labels, auto_labels, label)
        for label in LABELS
    ]

    macro_precision = sum(item["precision"] for item in metrics) / len(LABELS)
    macro_recall = sum(item["recall"] for item in metrics) / len(LABELS)
    macro_f1 = sum(item["f1"] for item in metrics) / len(LABELS)

    confusion_matrix = build_confusion_matrix(gold_labels, auto_labels)
    confusion_matrices[band] = confusion_matrix

    for item in metrics:
        all_metric_rows.append({
            "entropy_band": band,
            **item,
        })

    summary_rows.append({
        "entropy_band": band,
        "total_cases": total_cases,
        "correct_cases": correct_cases,
        "incorrect_cases": total_cases - correct_cases,
        "accuracy": accuracy,
        "macro_precision": macro_precision,
        "macro_recall": macro_recall,
        "macro_f1": macro_f1,
    })


# =========================
# CREAR EXCEL DE SALIDA
# =========================

wb_output = Workbook()


# Hoja 1: métricas por etiqueta y banda
ws_metrics = wb_output.active
ws_metrics.title = "per_band_label_metrics"

headers = [
    "entropy_band",
    "label",
    "gold_support",
    "auto_predicted",
    "true_positives",
    "false_positives",
    "false_negatives",
    "true_negatives",
    "precision",
    "recall",
    "f1",
]

ws_metrics.append(headers)
style_header(ws_metrics[1])

for item in all_metric_rows:
    ws_metrics.append([
        item["entropy_band"],
        item["label"],
        item["gold_support"],
        item["auto_predicted"],
        item["true_positives"],
        item["false_positives"],
        item["false_negatives"],
        item["true_negatives"],
        item["precision"],
        item["recall"],
        item["f1"],
    ])

for row in range(2, ws_metrics.max_row + 1):
    for col in ["I", "J", "K"]:
        ws_metrics[f"{col}{row}"].number_format = "0.0000"

apply_basic_style(ws_metrics)


# Hoja 2: resumen por banda
ws_summary = wb_output.create_sheet("summary_by_entropy_band")

summary_headers = [
    "entropy_band",
    "total_cases",
    "correct_cases",
    "incorrect_cases",
    "accuracy",
    "macro_precision",
    "macro_recall",
    "macro_f1",
]

ws_summary.append(summary_headers)
style_header(ws_summary[1])

for item in summary_rows:
    ws_summary.append([
        item["entropy_band"],
        item["total_cases"],
        item["correct_cases"],
        item["incorrect_cases"],
        item["accuracy"],
        item["macro_precision"],
        item["macro_recall"],
        item["macro_f1"],
    ])

for row in range(2, ws_summary.max_row + 1):
    for col in ["E", "F", "G", "H"]:
        ws_summary[f"{col}{row}"].number_format = "0.0000"

apply_basic_style(ws_summary)


# Hoja 3: matrices de confusión
ws_confusion = wb_output.create_sheet("confusion_matrices")

current_row = 1

for band in ENTROPY_BANDS:
    ws_confusion.cell(row=current_row, column=1).value = f"entropy_band = {band}"
    ws_confusion.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1

    ws_confusion.append(["Gold \\ Auto"] + LABELS)
    style_header(ws_confusion[current_row])
    current_row += 1

    matrix = confusion_matrices[band]

    for gold_label in LABELS:
        ws_confusion.append(
            [gold_label] + [
                matrix[gold_label][auto_label]
                for auto_label in LABELS
            ]
        )
        current_row += 1

    current_row += 2

apply_basic_style(ws_confusion)


# Hoja 4: filas inválidas
ws_invalid = wb_output.create_sheet("invalid_rows")

ws_invalid.append([
    "excel_row",
    "auto_label_raw",
    "gold_label_raw",
    "entropy_band_raw",
])
style_header(ws_invalid[1])

for item in invalid_rows:
    ws_invalid.append([
        item["excel_row"],
        item["auto_label_raw"],
        item["gold_label_raw"],
        item["entropy_band_raw"],
    ])

apply_basic_style(ws_invalid)


# Hoja 5: información metodológica
ws_info = wb_output.create_sheet("method_info")

method_rows = [
    ["input_file", str(INPUT_FILE)],
    ["input_sheet", ws_input.title],
    ["rows_checked", f"{FIRST_DATA_ROW}-{LAST_DATA_ROW}"],
    ["auto_label_column", AUTO_LABEL_COL],
    ["gold_label_column", GOLD_LABEL_COL],
    ["entropy_band_column", ENTROPY_BAND_COL],
    ["labels", ", ".join(LABELS)],
    ["entropy_bands", ", ".join(ENTROPY_BANDS)],
    ["evaluation_reference", "Gold Human Label"],
    ["evaluated_prediction", "auto_label"],
    ["macro_f1_definition", "Mean of F1_NEG, F1_NEU and F1_POS within each entropy band"],
]

for row in method_rows:
    ws_info.append(row)

apply_basic_style(ws_info)


# =========================
# GUARDAR EXCEL
# =========================

wb_output.save(OUTPUT_FILE)


# =========================
# RESULTADOS EN CONSOLA
# =========================

print("PRECISION, RECALL, F1 Y MACRO-F1 POR ENTROPY_BAND")
print("=================================================")
print(f"Archivo analizado: {INPUT_FILE}")
print(f"Hoja analizada: {ws_input.title}")
print(f"Filas analizadas: {FIRST_DATA_ROW}-{LAST_DATA_ROW}")
print()

for item in summary_rows:
    print(
        f"{item['entropy_band']}: "
        f"n={item['total_cases']}, "
        f"accuracy={item['accuracy']:.4f}, "
        f"macro-F1={item['macro_f1']:.4f}"
    )

if invalid_rows:
    print()
    print(f"ADVERTENCIA: {len(invalid_rows)} filas con valores inválidos o vacíos.")
    print("Primeros casos:")
    for item in invalid_rows[:20]:
        print(item)

print()
print(f"Archivo creado: {OUTPUT_FILE}")