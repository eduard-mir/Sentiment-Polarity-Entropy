from pathlib import Path
from openpyxl import load_workbook


# =========================
# CONFIGURACIÓN
# =========================

INPUT_FILE = Path("sample_1000_sentences_for_manual_annotation_randomized_auto_label_gold_entropy_band.xlsx")

SHEET_NAME = None  # None = usa la hoja activa

FIRST_DATA_ROW = 2
LAST_DATA_ROW = 960

AUTO_LABEL_COL = "I"
GOLD_LABEL_COL = "O"

VALID_LABELS = {"NEG", "NEU", "POS"}
EVALUATED_AUTO_LABELS = {"NEG", "POS"}


# =========================
# FUNCIONES
# =========================

def clean_label(value):
    if value is None:
        return None

    value = str(value).strip().upper()

    if value in VALID_LABELS:
        return value

    return None


def percentage(count, total):
    return (count / total) * 100 if total > 0 else 0


# =========================
# LECTURA DEL EXCEL
# =========================

wb = load_workbook(INPUT_FILE, data_only=True)

if SHEET_NAME is None:
    ws = wb.active
else:
    ws = wb[SHEET_NAME]


# =========================
# CÁLCULO DE ACCURACY
# SOLO AUTO_LABEL = POS / NEG
# =========================

total_rows_checked = 0
included_cases = 0
excluded_neu_cases = 0
invalid_or_missing_cases = 0

correct_cases = 0
incorrect_cases = 0

for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):

    total_rows_checked += 1

    auto_label = clean_label(ws[f"{AUTO_LABEL_COL}{row}"].value)
    gold_label = clean_label(ws[f"{GOLD_LABEL_COL}{row}"].value)

    if auto_label is None or gold_label is None:
        invalid_or_missing_cases += 1
        continue

    # Excluimos todos los casos en los que el sistema predijo NEU
    if auto_label == "NEU":
        excluded_neu_cases += 1
        continue

    # Evaluamos solo los casos en los que el sistema predijo NEG o POS
    if auto_label in EVALUATED_AUTO_LABELS:
        included_cases += 1

        if auto_label == gold_label:
            correct_cases += 1
        else:
            incorrect_cases += 1


accuracy = percentage(correct_cases, included_cases)


# =========================
# RESULTADOS EN CONSOLA
# =========================

print("ACCURACY AUTO_LABEL POS/NEG ONLY")
print("================================")
print(f"Archivo analizado: {INPUT_FILE}")
print(f"Hoja analizada: {ws.title}")
print(f"Filas analizadas: {FIRST_DATA_ROW}-{LAST_DATA_ROW}")
print()
print(f"Total de filas revisadas: {total_rows_checked}")
print(f"Casos incluidos en la evaluación: {included_cases}")
print(f"Casos excluidos porque auto_label = NEU: {excluded_neu_cases}")
print(f"Casos inválidos o vacíos: {invalid_or_missing_cases}")
print()
print(f"Casos correctos: {correct_cases}")
print(f"Casos incorrectos: {incorrect_cases}")
print(f"Accuracy POS/NEG only: {accuracy:.2f}%")